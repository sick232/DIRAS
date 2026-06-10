"""
Excel logger for DIRAS RAG pipeline

Provides `save_query_log_to_excel` which appends RAG query results
to `data/generated_answers/rag_debug_logs.xlsx` in sheet `RAG_Logs`.

This module is defensive: failures are logged and do not raise.
"""
from pathlib import Path
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


def _ensure_workbook(path: Path):
    from openpyxl import Workbook, load_workbook

    if path.exists():
        wb = load_workbook(path)
        if "RAG_Logs" not in wb.sheetnames:
            wb.create_sheet("RAG_Logs")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "RAG_Logs"
        headers = [
            "Timestamp",
            "Query",
            "Retrieved Chunks",
            "Retrieved Context",
            "Prompt Sent",
            "Gemini Output",
            "Groq Output",
            "Grok Output",
            "Fallback Output",
            "Final Answer",
            "Model Used",
            "Confidence",
            "Processing Time",
        ]
        ws.append(headers)
        from openpyxl.styles import Font
        for idx, _ in enumerate(headers, start=1):
            ws.cell(row=1, column=idx).font = Font(bold=True)
        ws.freeze_panes = "A2"

    return wb


def _auto_adjust_columns(ws):
    from openpyxl.utils import get_column_letter

    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column
        column_letter = get_column_letter(column)
        for cell in column_cells:
            try:
                value = cell.value or ""
                ln = len(str(value))
            except Exception:
                ln = 0
            if ln > max_length:
                max_length = ln
        adjusted_width = min(max_length + 5, 120)
        ws.column_dimensions[column_letter].width = adjusted_width


def _format_retrieved_chunks(retrieved_chunks):
    # If caller provided a pre-formatted string, return it unchanged
    if isinstance(retrieved_chunks, str):
        return retrieved_chunks

    parts = []
    for chunk in (retrieved_chunks or [])[:10]:
        title = None
        similarity = None
        text = None
        if isinstance(chunk, dict):
            metadata = chunk.get('metadata', {})
            title = metadata.get('document_title') or metadata.get('title') or metadata.get('document')
            similarity = chunk.get('similarity_score') or chunk.get('similarity')
            text = chunk.get('text') or chunk.get('full_text') or ''
        else:
            # fallback representation
            title = getattr(chunk, 'title', str(chunk))
            similarity = getattr(chunk, 'similarity', '')
            text = getattr(chunk, 'text', '')

        parts.append(
            f"Document: {title or 'Unknown'}\n"
            f"Similarity: {similarity if similarity is not None else ''}\n\n"
            f"Text:\n{text}"
        )

    return "\n\n---------------------------------------\n\n".join(parts)


def save_query_log_to_excel(
    query,
    retrieved_chunks,
    context,
    prompt_sent,
    gemini_output,
    groq_output,
    grok_output,
    fallback_output,
    final_answer,
    model_used,
    confidence,
    processing_time,
):
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment

        repo_root = Path(__file__).resolve().parents[2]
        log_folder = repo_root / "data" / "generated_answers"
        log_folder.mkdir(parents=True, exist_ok=True)
        log_path = log_folder / "rag_debug_logs.xlsx"

        # Ensure workbook/sheet exists
        wb = _ensure_workbook(log_path)

        # Save (ensures sheet exists and headers)
        if "RAG_Logs" in wb.sheetnames:
            ws = wb["RAG_Logs"]
        else:
            ws = wb.active
            ws.title = "RAG_Logs"

        timestamp = datetime.utcnow().isoformat()
        retrieved_chunks_text = _format_retrieved_chunks(retrieved_chunks)

        row = [
            timestamp,
            query,
            retrieved_chunks_text,
            context,
            prompt_sent,
            gemini_output,
            groq_output,
            grok_output,
            fallback_output,
            final_answer,
            model_used,
            confidence,
            processing_time,
        ]

        ws.append(row)

        # Formatting: wrap text, left align, vertical top
        for r in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in r:
                cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')

        _auto_adjust_columns(ws)

        wb.save(log_path)

    except Exception as e:
        logger.warning(f"Failed to save RAG debug log to Excel: {e}")
        return


def get_cached_query_result(query: str):
    """Return cached row for exact query match, or None if not found."""
    try:
        from openpyxl import load_workbook

        repo_root = Path(__file__).resolve().parents[2]
        log_path = repo_root / "data" / "generated_answers" / "rag_debug_logs.xlsx"
        if not log_path.exists():
            return None

        wb = load_workbook(log_path, read_only=True)
        if "RAG_Logs" not in wb.sheetnames:
            return None
        ws = wb["RAG_Logs"]

        # Find header indices
        headers = {cell.value: idx + 1 for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))}

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_query = row[headers.get('Query') - 1] if headers.get('Query') else None
            if row_query and str(row_query).strip() == str(query).strip():
                # map fields
                def _get(name):
                    idx = headers.get(name)
                    return row[idx - 1] if idx and idx - 1 < len(row) else None

                cached = {
                    'query': query,
                    'debug_chunks': _get('Retrieved Chunks') or '',
                    'retrieved_context': _get('Retrieved Context') or '',
                    'prompt_sent': _get('Prompt Sent') or '',
                    'gemini_raw_response': _get('Gemini Output') or '',
                    'groq_raw_response': _get('Groq Output') or '',
                    'grok_raw_response': _get('Grok Output') or '',
                    'fallback_raw_response': _get('Fallback Output') or '',
                    'final_answer': _get('Final Answer') or '',
                    'llm_outputs': {
                        'gemini': _get('Gemini Output') or '',
                        'groq': _get('Groq Output') or '',
                        'grok': _get('Grok Output') or '',
                        'fallback': _get('Fallback Output') or ''
                    },
                    'sources': [],
                    'processing_time': _get('Processing Time') or 0.0,
                    'model': _get('Model Used') or '',
                    'model_used': _get('Model Used') or '',
                    'confidence': _get('Confidence') or 0.0,
                    'answer': _get('Final Answer') or _get('Gemini Output') or ''
                }
                return cached

        return None
    except Exception as e:
        logger.warning(f"Failed to read cache from Excel: {e}")
        return None
